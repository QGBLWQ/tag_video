from pathlib import Path

from openpyxl import Workbook, load_workbook

from video_tagging_assistant.excel_models import ReviewSheetRow
from video_tagging_assistant.excel_workbook import load_confirmed_cases, sync_approved_rows, upsert_review_rows


def build_source_workbook(path: Path) -> None:
    wb = Workbook()
    source = wb.active
    source.title = "创建记录"
    source.append([
        "序号",
        "文件夹名",
        "备注",
        "Raw存放路径",
        "VS_Nomal",
        "VS_Night",
        "安装方式",
        "运动模式",
        "标签审核状态",
        "最终简介",
        "最终标签",
    ])
    source.append([
        1,
        "case_A_0001",
        "场景备注",
        "raw/path",
        "videos/case_A_0001/clip01.mp4",
        "videos/case_A_0001/night.mp4",
        "手持",
        "行走",
        "待生成",
        "",
        "",
    ])
    wb.save(path)


def test_load_confirmed_cases_reads_source_sheet(tmp_path: Path):
    workbook_path = tmp_path / "records.xlsx"
    build_source_workbook(workbook_path)

    rows = load_confirmed_cases(
        workbook_path,
        source_sheet="创建记录",
        case_key_column="文件夹名",
        status_column="标签审核状态",
    )

    assert len(rows) == 1
    assert rows[0].case_key == "case_A_0001"
    assert rows[0].raw_path == "raw/path"
    assert rows[0].vs_normal_path == "videos/case_A_0001/clip01.mp4"


def test_upsert_review_rows_creates_review_sheet(tmp_path: Path):
    workbook_path = tmp_path / "records.xlsx"
    build_source_workbook(workbook_path)

    upsert_review_rows(
        workbook_path,
        review_sheet="标签审核",
        rows=[
            ReviewSheetRow(
                case_key="case_A_0001",
                workbook_row_index=2,
                raw_path="raw/path",
                video_path="videos/case_A_0001/clip01.mp4",
                auto_summary="自动简介",
                auto_tags="安装方式=手持;运动模式=行走",
                auto_scene_description="画面描述",
            )
        ],
    )

    sheet = load_workbook(workbook_path)["标签审核"]
    assert sheet["A2"].value == "case_A_0001"
    assert sheet["E2"].value == "自动简介"
    assert sheet["F2"].value == "安装方式=手持;运动模式=行走"


def test_sync_approved_rows_updates_source_sheet(tmp_path: Path):
    workbook_path = tmp_path / "records.xlsx"
    build_source_workbook(workbook_path)
    upsert_review_rows(
        workbook_path,
        review_sheet="标签审核",
        rows=[
            ReviewSheetRow(
                case_key="case_A_0001",
                workbook_row_index=2,
                raw_path="raw/path",
                video_path="videos/case_A_0001/clip01.mp4",
                auto_summary="自动简介",
                auto_tags="安装方式=手持;运动模式=行走",
                auto_scene_description="画面描述",
                review_decision="审核通过",
            )
        ],
    )

    sync_approved_rows(workbook_path, source_sheet="创建记录", review_sheet="标签审核")

    source = load_workbook(workbook_path)["创建记录"]
    assert source["I2"].value == "审核通过"
    assert source["J2"].value == "自动简介"
    assert source["K2"].value == "安装方式=手持;运动模式=行走"
