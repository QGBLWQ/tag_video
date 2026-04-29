from pathlib import Path

import pytest
from openpyxl import Workbook
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QApplication

from video_tagging_assistant.gui import app as gui_app
from video_tagging_assistant.gui.main_window import PipelineMainWindow
from video_tagging_assistant.gui.review_panel import ReviewPanel
from video_tagging_assistant.gui.table_models import CaseTableModel
from video_tagging_assistant.pipeline_models import CaseManifest
from video_tagging_assistant.tagging_service import TaggingReviewRow


def test_pipeline_main_window_builds():
    app = QApplication.instance() or QApplication([])
    window = PipelineMainWindow()
    assert window.windowTitle() == "Case Pipeline"
    assert window.tabs.count() >= 3


def test_gui_exposes_tagging_mode_selector():
    app = QApplication.instance() or QApplication([])
    window = PipelineMainWindow()
    assert window.tagging_mode_combo.count() == 2
    assert window.tagging_mode_combo.itemText(0) == "重新打标"
    assert window.tagging_mode_combo.itemText(1) == "复用旧打标结果"


def test_gui_log_panel_appends_event_text():
    app = QApplication.instance() or QApplication([])
    window = PipelineMainWindow()
    window.append_log_line("case_A_0105 upload started")
    assert "upload started" in window.log_panel.toPlainText()


def test_review_panel_loads_row_and_collects_user_edits():
    app = QApplication.instance() or QApplication([])
    panel = ReviewPanel()
    panel.set_review_row(
        TaggingReviewRow(
            case_id="case_A_0105",
            auto_summary="自动简介",
            auto_tags="安装方式=手持",
            auto_scene_description="自动画面描述",
            tag_source="fresh",
        )
    )
    panel.manual_summary_edit.setPlainText("人工简介")
    panel.manual_tags_edit.setPlainText("安装方式=肩扛")
    payload = panel.current_review_payload()

    assert payload["case_id"] == "case_A_0105"
    assert payload["manual_summary"] == "人工简介"
    assert payload["manual_tags"] == "安装方式=肩扛"


def test_review_panel_refresh_button_calls_callback():
    app = QApplication.instance() or QApplication([])
    called = []
    panel = ReviewPanel(on_refresh_excel_reviews=lambda: called.append(True))

    panel.refresh_button.click()

    assert called == [True]


def test_case_table_model_displays_case_stage_and_message():
    model = CaseTableModel(
        [
            {
                "case_id": "case_A_0105",
                "stage": "awaiting_review",
                "tag_source": "fresh",
                "message": "awaiting review",
            }
        ]
    )

    assert model.rowCount() == 1
    assert model.columnCount() == 4
    assert model.data(model.index(0, 0), Qt.DisplayRole) == "case_A_0105"
    assert model.data(model.index(0, 1), Qt.DisplayRole) == "awaiting_review"


def test_main_window_scan_loads_cases_into_queue(tmp_path: Path):
    app = QApplication.instance() or QApplication([])
    called = []

    def fake_scan():
        called.append(True)
        return [
            CaseManifest(
                case_id="case_A_0105",
                row_index=2,
                created_date="20260428",
                mode="OV50H40_Action5Pro_DCG HDR",
                raw_path=tmp_path / "raw",
                vs_normal_path=tmp_path / "normal.MP4",
                vs_night_path=tmp_path / "night.MP4",
                local_case_root=tmp_path / "local" / "case_A_0105",
                server_case_dir=tmp_path / "server" / "case_A_0105",
                remark="场景备注",
                labels={"安装方式": "手持"},
            )
        ]

    window = PipelineMainWindow(scan_cases=fake_scan)
    window.scan_button.click()

    assert called == [True]
    assert window.queue_model.rowCount() == 1
    assert window.log_panel.toPlainText().strip().endswith("Scanned 1 cases")


def test_main_window_start_pipeline_loads_review_panel(tmp_path: Path):
    app = QApplication.instance() or QApplication([])
    manifest = CaseManifest(
        case_id="case_A_0105",
        row_index=2,
        created_date="20260428",
        mode="OV50H40_Action5Pro_DCG HDR",
        raw_path=tmp_path / "raw",
        vs_normal_path=tmp_path / "normal.MP4",
        vs_night_path=tmp_path / "night.MP4",
        local_case_root=tmp_path / "local" / "case_A_0105",
        server_case_dir=tmp_path / "server" / "case_A_0105",
        remark="场景备注",
        labels={"安装方式": "手持"},
    )
    captured = []

    def fake_start_tagging(manifests, mode, event_callback):
        captured.append((manifests[0].case_id, mode))
        event_callback(
            type(
                "Event",
                (),
                {
                    "case_id": "case_A_0105",
                    "stage": type("Stage", (), {"value": "tagging_running"})(),
                    "message": "tagging",
                },
            )()
        )
        return [
            TaggingReviewRow(
                case_id="case_A_0105",
                auto_summary="自动简介",
                auto_tags="安装方式=手持",
                auto_scene_description="自动画面描述",
                tag_source="fresh",
            )
        ]

    window = PipelineMainWindow(start_tagging=fake_start_tagging)
    window._manifests_by_case_id = {manifest.case_id: manifest}
    window.start_button.click()

    assert captured == [("case_A_0105", "fresh")]
    assert window.review_panel.case_label.text() == "case_A_0105"
    assert "tagging" in window.log_panel.toPlainText()


def test_gui_approve_calls_controller_and_execution_runner(tmp_path: Path):
    app = QApplication.instance() or QApplication([])
    approvals = []
    runs = []

    class StubController:
        def approve_case(self, case_id):
            approvals.append(case_id)
            return True

    window = PipelineMainWindow(run_execution_case=lambda case_id: runs.append(case_id), controller=StubController())
    window._review_rows_by_case_id = {
        "case_A_0105": TaggingReviewRow(
            case_id="case_A_0105",
            auto_summary="自动简介",
            auto_tags="安装方式=手持",
            auto_scene_description="自动画面描述",
            tag_source="fresh",
        )
    }
    window.review_panel.set_review_row(window._review_rows_by_case_id["case_A_0105"])

    window.review_panel.approve_button.click()

    assert approvals == ["case_A_0105"]
    assert runs == ["case_A_0105"]


def test_refresh_excel_reviews_only_runs_newly_approved_cases():
    app = QApplication.instance() or QApplication([])
    approvals = []
    runs = []

    class StubController:
        def approve_case(self, case_id):
            approvals.append(case_id)
            if case_id == "case_A_0105":
                return True
            return False

    window = PipelineMainWindow(
        controller=StubController(),
        run_execution_case=lambda case_id: runs.append(case_id),
        refresh_excel_reviews=lambda: [
            {"case_id": "case_A_0105", "review_decision": "审核通过", "manual_summary": "", "manual_tags": "", "review_note": ""},
            {"case_id": "case_A_0106", "review_decision": "审核通过", "manual_summary": "", "manual_tags": "", "review_note": ""},
        ],
    )

    window.review_panel.refresh_button.click()

    assert approvals == ["case_A_0105", "case_A_0106"]
    assert runs == ["case_A_0105"]


def test_launch_case_pipeline_gui_passes_workbook_path(monkeypatch, tmp_path: Path):
    captured = {}

    class FakeWindow:
        def __init__(self, workbook_path=None, **kwargs):
            captured["workbook_path"] = workbook_path

        def show(self):
            captured["shown"] = True

    class FakeQApplication:
        @staticmethod
        def instance():
            return None

        def __init__(self, argv):
            pass

        def exec_(self):
            return 0

    monkeypatch.setattr(gui_app, "PipelineMainWindow", FakeWindow)
    monkeypatch.setattr(gui_app, "QApplication", FakeQApplication)

    gui_app.launch_case_pipeline_gui(workbook_path=str(tmp_path / "records.xlsx"))

    assert captured["workbook_path"].endswith("records.xlsx")
    assert captured["shown"] is True


def test_launch_case_pipeline_gui_injects_real_scan_and_refresh(monkeypatch, tmp_path: Path):
    workbook_path = tmp_path / "records.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "创建记录"
    ws.append(["序号", "文件夹名", "备注", "创建日期", "Raw存放路径", "VS_Nomal", "VS_Night", "安装方式", "运动模式"])
    ws.append([1, "case_A_0105", "场景备注", "20260428", r"E:\DV\raw", r"E:\DV\normal.MP4", r"E:\DV\night.MP4", "手持", "行走"])
    review = wb.create_sheet("审核结果")
    review.append(["文件夹名", "创建记录行号", "Raw存放路径", "视频路径", "自动简介", "自动标签", "自动画面描述", "审核结论", "人工修订简介", "人工修订标签", "审核备注", "审核人", "审核时间", "同步状态", "归档状态", "归档目标路径"])
    wb.save(workbook_path)

    captured = {}

    class FakeWindow:
        def __init__(self, workbook_path=None, scan_cases=None, refresh_excel_reviews=None, **kwargs):
            captured["workbook_path"] = workbook_path
            captured["scan_cases"] = scan_cases
            captured["refresh_excel_reviews"] = refresh_excel_reviews

        def show(self):
            captured["shown"] = True

    class FakeQApplication:
        @staticmethod
        def instance():
            return None

        def __init__(self, argv):
            pass

        def exec_(self):
            return 0

    monkeypatch.setattr(gui_app, "PipelineMainWindow", FakeWindow)
    monkeypatch.setattr(gui_app, "QApplication", FakeQApplication)
    monkeypatch.setattr(
        gui_app,
        "load_config",
        lambda path: {
            "input_dir": "videos",
            "output_dir": "output",
            "compression": {},
            "provider": {"name": "mock", "model": "mock-model"},
            "prompt_template": {"system": "describe"},
            "gui_pipeline": {
                "source_sheet": "创建记录",
                "review_sheet": "审核结果",
            },
        },
    )

    gui_app.launch_case_pipeline_gui(workbook_path=str(workbook_path))

    manifests = captured["scan_cases"]()
    approved = captured["refresh_excel_reviews"]()

    assert captured["workbook_path"].endswith("records.xlsx")
    assert len(manifests) == 1
    assert manifests[0].case_id == "case_A_0105"
    assert approved == []
    assert captured["shown"] is True


def test_launch_case_pipeline_gui_injects_execution_bridge(monkeypatch, tmp_path: Path):
    captured = {}
    execution_calls = []

    class FakeController:
        def has_execution_case(self):
            execution_calls.append("has")
            return True

        def run_next_execution_case(self):
            execution_calls.append("run")

    class FakeWindow:
        def __init__(self, run_execution_case=None, **kwargs):
            captured["run_execution_case"] = run_execution_case

        def show(self):
            pass

    class FakeQApplication:
        @staticmethod
        def instance():
            return None

        def __init__(self, argv):
            pass

        def exec_(self):
            return 0

    monkeypatch.setattr(gui_app, "PipelineController", lambda **kwargs: FakeController())
    monkeypatch.setattr(gui_app, "PipelineMainWindow", FakeWindow)
    monkeypatch.setattr(gui_app, "QApplication", FakeQApplication)

    gui_app.launch_case_pipeline_gui(workbook_path=str(tmp_path / "records.xlsx"))
    captured["run_execution_case"]("case_A_0105")

    assert execution_calls == ["has", "run"]


def test_launch_case_pipeline_gui_enters_event_loop(monkeypatch, tmp_path: Path):
    captured = {"exec_calls": 0}

    class FakeApp:
        @staticmethod
        def instance():
            return None

        def __init__(self, argv):
            pass

        def exec_(self):
            captured["exec_calls"] += 1
            return 0

    class FakeWindow:
        def __init__(self, **kwargs):
            pass

        def show(self):
            captured["shown"] = True

    monkeypatch.setattr(gui_app, "QApplication", FakeApp)
    monkeypatch.setattr(gui_app, "PipelineMainWindow", FakeWindow)

    result = gui_app.launch_case_pipeline_gui(workbook_path=str(tmp_path / "records.xlsx"))

    assert captured["shown"] is True
    assert captured["exec_calls"] == 1
    assert result == 0


def test_launch_case_pipeline_gui_injects_tagging_bridge(monkeypatch, tmp_path: Path):
    captured = {}
    provider_calls = []
    tagging_calls = []

    class FakeWindow:
        def __init__(self, start_tagging=None, **kwargs):
            captured["start_tagging"] = start_tagging

        def show(self):
            pass

    class FakeQApplication:
        @staticmethod
        def instance():
            return None

        def __init__(self, argv):
            pass

        def exec_(self):
            return 0

    monkeypatch.setattr(gui_app, "PipelineMainWindow", FakeWindow)
    monkeypatch.setattr(gui_app, "QApplication", FakeQApplication)
    monkeypatch.setattr(
        gui_app,
        "load_config",
        lambda path: {
            "prompt_template": {"system": "describe"},
            "provider": {"name": "mock", "model": "mock-model"},
            "input_dir": "in",
            "output_dir": "out",
            "compression": {},
        },
    )
    monkeypatch.setattr(gui_app, "build_provider_from_config", lambda config: provider_calls.append(config["provider"]) or object())
    monkeypatch.setattr(
        gui_app,
        "run_batch_tagging",
        lambda manifests, cache_root, output_root, provider, prompt_template, mode, event_callback: tagging_calls.append(
            {
                "count": len(manifests),
                "mode": mode,
                "prompt_template": prompt_template,
                "cache_root": cache_root,
                "output_root": output_root,
                "provider": provider,
            }
        ) or [],
    )

    gui_app.launch_case_pipeline_gui(workbook_path=str(tmp_path / "records.xlsx"))

    manifest = CaseManifest(
        case_id="case_A_0105",
        row_index=2,
        created_date="20260428",
        mode="OV50H40_Action5Pro_DCG HDR",
        raw_path=tmp_path / "raw",
        vs_normal_path=tmp_path / "normal.MP4",
        vs_night_path=tmp_path / "night.MP4",
        local_case_root=tmp_path / "local" / "case_A_0105",
        server_case_dir=tmp_path / "server" / "case_A_0105",
        remark="场景备注",
        labels={"安装方式": "手持"},
    )

    captured["start_tagging"]([manifest], "fresh", lambda event: None)

    assert provider_calls == [{"name": "mock", "model": "mock-model"}]
    assert tagging_calls == [
        {
            "count": 1,
            "mode": "fresh",
            "prompt_template": {"system": "describe"},
            "cache_root": gui_app.DEFAULT_CACHE_ROOT,
            "output_root": gui_app.DEFAULT_TAGGING_OUTPUT_ROOT,
            "provider": tagging_calls[0]["provider"],
        }
    ]


def test_launch_case_pipeline_gui_bridges_get_list_into_manifests(monkeypatch, tmp_path: Path):
    workbook_path = tmp_path / "records.xlsx"
    wb = Workbook()
    create_record = wb.active
    create_record.title = "创建记录"
    create_record.append(["序号", "文件夹名", "备注", "创建日期", "Raw存放路径", "VS_Nomal", "VS_Night", "安装方式", "运动模式", "pipeline_status"])
    create_record.append([
        1,
        "case_A_0001",
        "场景备注",
        "20260422",
        r"E:\DV\case_A_0001\case_A_0001_RK_raw_117",
        r"E:\DV\case_A_0001\DJI_20260422151829_0001_D.MP4",
        r"E:\DV\case_A_0001\DJI_20260422151916_0021_D.MP4",
        "手持",
        "行走",
        "queued",
    ])
    get_list = wb.create_sheet("获取列表")
    get_list.append(["日期", "20260422", "", ""])
    get_list.append(["处理状态", "RK_raw", "Action5Pro_Nomal", "Action5Pro_Night"])
    get_list.append(["R", "117", "DJI_20260422151829_0001_D.MP4", "DJI_20260422151916_0021_D.MP4"])
    review = wb.create_sheet("审核结果")
    review.append(["文件夹名", "创建记录行号", "Raw存放路径", "视频路径", "自动简介", "自动标签", "自动画面描述", "审核结论", "人工修订简介", "人工修订标签", "审核备注", "审核人", "审核时间", "同步状态", "归档状态", "归档目标路径"])
    wb.save(workbook_path)

    captured = {}

    class FakeWindow:
        def __init__(self, scan_cases=None, **kwargs):
            captured["scan_cases"] = scan_cases

        def show(self):
            pass

    class FakeQApplication:
        @staticmethod
        def instance():
            return None

        def __init__(self, argv):
            pass

        def exec_(self):
            return 0

    monkeypatch.setattr(gui_app, "PipelineMainWindow", FakeWindow)
    monkeypatch.setattr(gui_app, "QApplication", FakeQApplication)
    monkeypatch.setattr(
        gui_app,
        "load_config",
        lambda path: {
            "input_dir": "videos",
            "output_dir": "output",
            "compression": {},
            "provider": {"name": "mock", "model": "mock-model"},
            "prompt_template": {"system": "describe"},
            "gui_pipeline": {
                "source_sheet": "获取列表",
                "review_sheet": "审核结果",
                "mode": "OV50H40_Action5Pro_DCG HDR",
                "allowed_statuses": ["queued"],
                "local_root": str(tmp_path / "local"),
                "server_root": str(tmp_path / "server"),
            },
        },
    )

    gui_app.launch_case_pipeline_gui(workbook_path=str(workbook_path))
    manifests = captured["scan_cases"]()

    assert [manifest.case_id for manifest in manifests] == ["case_A_0001"]



def test_launch_case_pipeline_gui_uses_create_record_source_sheet(monkeypatch, tmp_path: Path):
    workbook_path = tmp_path / "records.xlsx"
    workbook_path.write_text("placeholder", encoding="utf-8")
    captured = {}
    calls = {}

    class FakeWindow:
        def __init__(self, scan_cases=None, **kwargs):
            captured["scan_cases"] = scan_cases

        def show(self):
            pass

    class FakeQApplication:
        @staticmethod
        def instance():
            return None

        def __init__(self, argv):
            pass

        def exec_(self):
            return 0

    monkeypatch.setattr(gui_app, "PipelineMainWindow", FakeWindow)
    monkeypatch.setattr(gui_app, "QApplication", FakeQApplication)
    monkeypatch.setattr(
        gui_app,
        "load_config",
        lambda path: {
            "input_dir": "videos",
            "output_dir": "output",
            "compression": {},
            "provider": {"name": "mock", "model": "mock-model"},
            "prompt_template": {"system": "describe"},
            "gui_pipeline": {
                "source_sheet": "创建记录",
                "review_sheet": "审核结果",
            },
        },
    )
    monkeypatch.setattr(gui_app, "ensure_pipeline_columns", lambda workbook, source_sheet: calls.update({"source_sheet": source_sheet}))
    monkeypatch.setattr(
        gui_app,
        "build_case_manifests",
        lambda workbook, source_sheet, allowed_statuses, local_root, server_root, mode: calls.update({"manifest_source_sheet": source_sheet}) or [],
    )

    gui_app.launch_case_pipeline_gui(workbook_path=str(workbook_path))
    captured["scan_cases"]()

    assert calls["source_sheet"] == "创建记录"
    assert calls["manifest_source_sheet"] == "创建记录"



def test_launch_case_pipeline_gui_keeps_excel_tagging_inputs_in_excel_mode(monkeypatch, tmp_path: Path):
    captured = {}
    tagging_calls = []

    class FakeWindow:
        def __init__(self, start_tagging=None, **kwargs):
            captured["start_tagging"] = start_tagging

        def show(self):
            pass

    class FakeQApplication:
        @staticmethod
        def instance():
            return None

        def __init__(self, argv):
            pass

        def exec_(self):
            return 0

    monkeypatch.setattr(gui_app, "PipelineMainWindow", FakeWindow)
    monkeypatch.setattr(gui_app, "QApplication", FakeQApplication)
    monkeypatch.setattr(
        gui_app,
        "load_config",
        lambda path: {
            "input_dir": "videos",
            "output_dir": "output",
            "compression": {},
            "provider": {"name": "mock", "model": "mock-model"},
            "prompt_template": {"system": "describe"},
            "gui_pipeline": {
                "source_sheet": "创建记录",
                "review_sheet": "审核结果",
                "tagging_input_mode": "excel",
                "tagging_input_root": "ignored",
            },
        },
    )
    monkeypatch.setattr(gui_app, "build_provider_from_config", lambda config: object())
    monkeypatch.setattr(
        gui_app,
        "run_batch_tagging",
        lambda manifests, cache_root, output_root, provider, prompt_template, mode, event_callback: tagging_calls.append(manifests) or [],
    )

    gui_app.launch_case_pipeline_gui(workbook_path=str(tmp_path / "records.xlsx"))
    manifest = CaseManifest(
        case_id="case_A_0105",
        row_index=2,
        created_date="20260428",
        mode="OV50H40_Action5Pro_DCG HDR",
        raw_path=tmp_path / "raw",
        vs_normal_path=tmp_path / "excel_normal.MP4",
        vs_night_path=tmp_path / "excel_night.MP4",
        local_case_root=tmp_path / "local" / "case_A_0105",
        server_case_dir=tmp_path / "server" / "case_A_0105",
        remark="场景备注",
        labels={"安装方式": "手持"},
    )

    captured["start_tagging"]([manifest], "fresh", lambda event: None)

    assert tagging_calls == [[manifest]]


def test_launch_case_pipeline_gui_remaps_tagging_inputs_from_local_root(monkeypatch, tmp_path: Path):
    captured = {}
    provider_calls = []
    tagging_calls = []
    local_video_root = tmp_path / "videos"
    local_video_root.mkdir()
    local_normal = local_video_root / "normal.MP4"
    local_night = local_video_root / "night.MP4"
    local_normal.write_text("normal", encoding="utf-8")
    local_night.write_text("night", encoding="utf-8")

    class FakeWindow:
        def __init__(self, start_tagging=None, **kwargs):
            captured["start_tagging"] = start_tagging

        def show(self):
            pass

    class FakeQApplication:
        @staticmethod
        def instance():
            return None

        def __init__(self, argv):
            pass

        def exec_(self):
            return 0

    monkeypatch.setattr(gui_app, "PipelineMainWindow", FakeWindow)
    monkeypatch.setattr(gui_app, "QApplication", FakeQApplication)
    monkeypatch.setattr(
        gui_app,
        "load_config",
        lambda path: {
            "input_dir": "videos",
            "output_dir": "output",
            "compression": {},
            "provider": {"name": "mock", "model": "mock-model"},
            "prompt_template": {"system": "describe"},
            "gui_pipeline": {
                "tagging_input_mode": "local_root",
                "tagging_input_root": str(local_video_root),
                "cache_root": "my_cache",
                "tagging_output_root": "my_gui_output",
            },
        },
    )
    monkeypatch.setattr(gui_app, "build_provider_from_config", lambda config: provider_calls.append(config["provider"]) or object())
    monkeypatch.setattr(
        gui_app,
        "run_batch_tagging",
        lambda manifests, cache_root, output_root, provider, prompt_template, mode, event_callback: tagging_calls.append(
            {
                "manifests": manifests,
                "cache_root": cache_root,
                "output_root": output_root,
                "mode": mode,
            }
        ) or [],
    )

    gui_app.launch_case_pipeline_gui(workbook_path=str(tmp_path / "records.xlsx"))

    manifest = CaseManifest(
        case_id="case_A_0105",
        row_index=2,
        created_date="20260428",
        mode="OV50H40_Action5Pro_DCG HDR",
        raw_path=tmp_path / "raw",
        vs_normal_path=Path(r"\\10.10.10.164\rk3668_capture\normal.MP4"),
        vs_night_path=Path(r"\\10.10.10.164\rk3668_capture\night.MP4"),
        local_case_root=tmp_path / "local" / "case_A_0105",
        server_case_dir=tmp_path / "server" / "case_A_0105",
        remark="场景备注",
        labels={"安装方式": "手持"},
    )

    captured["start_tagging"]([manifest], "fresh", lambda event: None)

    assert provider_calls == [{"name": "mock", "model": "mock-model"}]
    assert tagging_calls[0]["cache_root"] == Path("my_cache")
    assert tagging_calls[0]["output_root"] == Path("my_gui_output")
    remapped_manifest = tagging_calls[0]["manifests"][0]
    assert remapped_manifest is not manifest
    assert remapped_manifest.case_id == "case_A_0105"
    assert remapped_manifest.vs_normal_path == local_normal
    assert remapped_manifest.vs_night_path == local_night
    assert manifest.vs_normal_path == Path(r"\\10.10.10.164\rk3668_capture\normal.MP4")



    class FakeWindow:
        def __init__(self, scan_cases=None, start_tagging=None, refresh_excel_reviews=None, **kwargs):
            captured["scan_cases"] = scan_cases
            captured["start_tagging"] = start_tagging
            captured["refresh_excel_reviews"] = refresh_excel_reviews

        def show(self):
            pass

    class FakeQApplication:
        @staticmethod
        def instance():
            return None

        def __init__(self, argv):
            pass

        def exec_(self):
            return 0

    monkeypatch.setattr(gui_app, "PipelineMainWindow", FakeWindow)
    monkeypatch.setattr(gui_app, "QApplication", FakeQApplication)
    monkeypatch.setattr(
        gui_app,
        "load_config",
        lambda path: {
            "input_dir": "videos",
            "output_dir": "output",
            "compression": {},
            "provider": {"name": "mock", "model": "mock-model"},
            "prompt_template": {"system": "describe"},
            "gui_pipeline": {
                "source_sheet": "我的创建记录",
                "review_sheet": "我的审核表",
                "mode": "MY_MODE",
                "allowed_statuses": ["pending", "failed"],
                "local_root": "my_cases",
                "server_root": "my_server_cases",
                "cache_root": "my_cache",
                "tagging_output_root": "my_gui_output",
            },
        },
    )

    calls = {}
    monkeypatch.setattr(gui_app, "ensure_pipeline_columns", lambda workbook, source_sheet: calls.update({"source_sheet": source_sheet}))
    monkeypatch.setattr(
        gui_app,
        "build_case_manifests",
        lambda workbook, source_sheet, allowed_statuses, local_root, server_root, mode: calls.update(
            {
                "manifest_source_sheet": source_sheet,
                "allowed_statuses": allowed_statuses,
                "local_root": local_root,
                "server_root": server_root,
                "mode": mode,
            }
        )
        or [],
    )
    monkeypatch.setattr(gui_app, "load_approved_review_rows", lambda workbook, review_sheet: calls.update({"review_sheet": review_sheet}) or [])
    monkeypatch.setattr(gui_app, "build_provider_from_config", lambda config: object())
    monkeypatch.setattr(
        gui_app,
        "run_batch_tagging",
        lambda manifests, cache_root, output_root, provider, prompt_template, mode, event_callback: calls.update(
            {"cache_root": cache_root, "output_root": output_root, "tag_mode": mode}
        )
        or [],
    )

    workbook_path = tmp_path / "records.xlsx"
    workbook_path.write_text("placeholder", encoding="utf-8")
    gui_app.launch_case_pipeline_gui(workbook_path=str(workbook_path))
    captured["scan_cases"]()
    captured["refresh_excel_reviews"]()
    captured["start_tagging"]([], "fresh", lambda event: None)

    assert calls["source_sheet"] == "我的创建记录"
    assert calls["manifest_source_sheet"] == "我的创建记录"
    assert calls["review_sheet"] == "我的审核表"
    assert calls["allowed_statuses"] == {"pending", "failed"}
    assert calls["local_root"] == Path("my_cases")
    assert calls["server_root"] == Path("my_server_cases")
    assert calls["mode"] == "MY_MODE"


def test_launch_case_pipeline_gui_rewrites_upload_target_when_local_upload_enabled(monkeypatch, tmp_path: Path):
    captured = {}
    upload_calls = []

    class FakeController:
        def __init__(self, pull_runner=None, copy_runner=None, upload_runner=None, event_callback=None):
            captured["upload_runner"] = upload_runner

        def has_execution_case(self):
            return False

        def run_next_execution_case(self):
            return None

    class FakeWindow:
        def __init__(self, **kwargs):
            pass

        def show(self):
            pass

    class FakeQApplication:
        @staticmethod
        def instance():
            return None

        def __init__(self, argv):
            pass

        def exec_(self):
            return 0

    monkeypatch.setattr(gui_app, "PipelineController", FakeController)
    monkeypatch.setattr(gui_app, "PipelineMainWindow", FakeWindow)
    monkeypatch.setattr(gui_app, "QApplication", FakeQApplication)
    monkeypatch.setattr(
        gui_app,
        "load_config",
        lambda path: {
            "input_dir": "videos",
            "output_dir": "output",
            "compression": {},
            "provider": {"name": "mock", "model": "mock-model"},
            "prompt_template": {"system": "describe"},
            "gui_pipeline": {
                "source_sheet": "创建记录",
                "review_sheet": "审核结果",
                "local_upload_enabled": True,
                "local_upload_root": str(tmp_path / "mock_server_cases"),
            },
        },
    )
    monkeypatch.setattr(
        gui_app,
        "upload_case_directory",
        lambda case_id, local_case_dir, server_case_dir, progress_callback=None: upload_calls.append(
            (case_id, local_case_dir, server_case_dir)
        ),
    )

    gui_app.launch_case_pipeline_gui(workbook_path=str(tmp_path / "records.xlsx"))
    captured["upload_runner"](
        "case_A_0001",
        tmp_path / "cases" / "OV50H40_Action5Pro_DCG HDR" / "20260414" / "case_A_0001",
        tmp_path / "server_cases" / "OV50H40_Action5Pro_DCG HDR" / "20260414" / "case_A_0001",
    )

    assert upload_calls == [
        (
            "case_A_0001",
            tmp_path / "cases" / "OV50H40_Action5Pro_DCG HDR" / "20260414" / "case_A_0001",
            tmp_path / "mock_server_cases" / "OV50H40_Action5Pro_DCG HDR" / "20260414" / "case_A_0001",
        )
    ]



def test_launch_case_pipeline_gui_keeps_server_upload_target_when_local_upload_disabled(monkeypatch, tmp_path: Path):
    captured = {}
    upload_calls = []

    class FakeController:
        def __init__(self, pull_runner=None, copy_runner=None, upload_runner=None, event_callback=None):
            captured["upload_runner"] = upload_runner

        def has_execution_case(self):
            return False

        def run_next_execution_case(self):
            return None

    class FakeWindow:
        def __init__(self, **kwargs):
            pass

        def show(self):
            pass

    class FakeQApplication:
        @staticmethod
        def instance():
            return None

        def __init__(self, argv):
            pass

        def exec_(self):
            return 0

    monkeypatch.setattr(gui_app, "PipelineController", FakeController)
    monkeypatch.setattr(gui_app, "PipelineMainWindow", FakeWindow)
    monkeypatch.setattr(gui_app, "QApplication", FakeQApplication)
    monkeypatch.setattr(
        gui_app,
        "load_config",
        lambda path: {
            "input_dir": "videos",
            "output_dir": "output",
            "compression": {},
            "provider": {"name": "mock", "model": "mock-model"},
            "prompt_template": {"system": "describe"},
            "gui_pipeline": {
                "source_sheet": "创建记录",
                "review_sheet": "审核结果",
                "local_upload_enabled": False,
                "local_upload_root": str(tmp_path / "mock_server_cases"),
            },
        },
    )
    monkeypatch.setattr(
        gui_app,
        "upload_case_directory",
        lambda case_id, local_case_dir, server_case_dir, progress_callback=None: upload_calls.append(
            (case_id, local_case_dir, server_case_dir)
        ),
    )

    gui_app.launch_case_pipeline_gui(workbook_path=str(tmp_path / "records.xlsx"))
    original_target = tmp_path / "server_cases" / "OV50H40_Action5Pro_DCG HDR" / "20260414" / "case_A_0001"
    captured["upload_runner"](
        "case_A_0001",
        tmp_path / "cases" / "OV50H40_Action5Pro_DCG HDR" / "20260414" / "case_A_0001",
        original_target,
    )

    assert upload_calls == [
        (
            "case_A_0001",
            tmp_path / "cases" / "OV50H40_Action5Pro_DCG HDR" / "20260414" / "case_A_0001",
            original_target,
        )
    ]



def test_launch_case_pipeline_gui_requires_local_upload_root_when_local_upload_enabled(monkeypatch, tmp_path: Path):
    class FakeWindow:
        def __init__(self, **kwargs):
            pass

        def show(self):
            pass

    class FakeQApplication:
        @staticmethod
        def instance():
            return None

        def __init__(self, argv):
            pass

        def exec_(self):
            return 0

    monkeypatch.setattr(gui_app, "PipelineMainWindow", FakeWindow)
    monkeypatch.setattr(gui_app, "QApplication", FakeQApplication)
    monkeypatch.setattr(
        gui_app,
        "load_config",
        lambda path: {
            "input_dir": "videos",
            "output_dir": "output",
            "compression": {},
            "provider": {"name": "mock", "model": "mock-model"},
            "prompt_template": {"system": "describe"},
            "gui_pipeline": {
                "source_sheet": "创建记录",
                "review_sheet": "审核结果",
                "local_upload_enabled": True,
                "local_upload_root": "",
            },
        },
    )

    with pytest.raises(ValueError) as exc:
        gui_app.launch_case_pipeline_gui(workbook_path=str(tmp_path / "records.xlsx"))

    assert "local_upload_root" in str(exc.value)


    captured = {}

    class FakeWindow:
        def __init__(self, start_tagging=None, **kwargs):
            captured["start_tagging"] = start_tagging

        def show(self):
            pass

    class FakeQApplication:
        @staticmethod
        def instance():
            return None

        def __init__(self, argv):
            pass

        def exec_(self):
            return 0

    monkeypatch.setattr(gui_app, "PipelineMainWindow", FakeWindow)
    monkeypatch.setattr(gui_app, "QApplication", FakeQApplication)
    monkeypatch.setattr(
        gui_app,
        "load_config",
        lambda path: {
            "input_dir": "videos",
            "output_dir": "output",
            "compression": {},
            "provider": {"name": "mock", "model": "mock-model"},
            "prompt_template": {"system": "describe"},
            "gui_pipeline": {
                "tagging_input_mode": "local_root",
                "tagging_input_root": str(tmp_path / "missing_videos"),
            },
        },
    )
    monkeypatch.setattr(gui_app, "build_provider_from_config", lambda config: object())

    gui_app.launch_case_pipeline_gui(workbook_path=str(tmp_path / "records.xlsx"))
    manifest = CaseManifest(
        case_id="case_A_0001",
        row_index=2,
        created_date="20260428",
        mode="OV50H40_Action5Pro_DCG HDR",
        raw_path=tmp_path / "raw",
        vs_normal_path=Path(r"\\10.10.10.164\rk3668_capture\case_A_0001_DJI_20260414144209_0109_D.DNG"),
        vs_night_path=Path(r"\\10.10.10.164\rk3668_capture\case_A_0001_DJI_20260414144209_0109_N.DNG"),
        local_case_root=tmp_path / "local" / "case_A_0001",
        server_case_dir=tmp_path / "server" / "case_A_0001",
        remark="场景备注",
        labels={"安装方式": "手持"},
    )

    with pytest.raises(FileNotFoundError) as exc:
        captured["start_tagging"]([manifest], "fresh", lambda event: None)

    assert "case_A_0001" in str(exc.value)
    assert "case_A_0001_DJI_20260414144209_0109_D.DNG" in str(exc.value)
    assert str(tmp_path / "missing_videos") in str(exc.value)



def test_launch_case_pipeline_gui_rejects_invalid_tagging_input_mode(monkeypatch, tmp_path: Path):
    captured = {}

    class FakeWindow:
        def __init__(self, start_tagging=None, **kwargs):
            captured["start_tagging"] = start_tagging

        def show(self):
            pass

    class FakeQApplication:
        @staticmethod
        def instance():
            return None

        def __init__(self, argv):
            pass

        def exec_(self):
            return 0

    monkeypatch.setattr(gui_app, "PipelineMainWindow", FakeWindow)
    monkeypatch.setattr(gui_app, "QApplication", FakeQApplication)
    monkeypatch.setattr(
        gui_app,
        "load_config",
        lambda path: {
            "input_dir": "videos",
            "output_dir": "output",
            "compression": {},
            "provider": {"name": "mock", "model": "mock-model"},
            "prompt_template": {"system": "describe"},
            "gui_pipeline": {
                "source_sheet": "创建记录",
                "review_sheet": "审核结果",
                "tagging_input_mode": "bad_mode",
                "tagging_input_root": "videos",
            },
        },
    )
    monkeypatch.setattr(gui_app, "build_provider_from_config", lambda config: object())

    gui_app.launch_case_pipeline_gui(workbook_path=str(tmp_path / "records.xlsx"))
    manifest = CaseManifest(
        case_id="case_A_0105",
        row_index=2,
        created_date="20260428",
        mode="OV50H40_Action5Pro_DCG HDR",
        raw_path=tmp_path / "raw",
        vs_normal_path=tmp_path / "excel_normal.MP4",
        vs_night_path=tmp_path / "excel_night.MP4",
        local_case_root=tmp_path / "local" / "case_A_0105",
        server_case_dir=tmp_path / "server" / "case_A_0105",
        remark="场景备注",
        labels={"安装方式": "手持"},
    )

    with pytest.raises(ValueError) as exc:
        captured["start_tagging"]([manifest], "fresh", lambda event: None)

    assert "tagging_input_mode" in str(exc.value)
    assert "bad_mode" in str(exc.value)


def test_launch_case_pipeline_gui_falls_back_when_gui_pipeline_missing(monkeypatch, tmp_path: Path):
    captured = {}

    class FakeWindow:
        def __init__(self, scan_cases=None, **kwargs):
            captured["scan_cases"] = scan_cases

        def show(self):
            pass

    class FakeQApplication:
        @staticmethod
        def instance():
            return None

        def __init__(self, argv):
            pass

        def exec_(self):
            return 0

    monkeypatch.setattr(gui_app, "PipelineMainWindow", FakeWindow)
    monkeypatch.setattr(gui_app, "QApplication", FakeQApplication)
    monkeypatch.setattr(
        gui_app,
        "load_config",
        lambda path: {
            "input_dir": "videos",
            "output_dir": "output",
            "compression": {},
            "provider": {"name": "mock", "model": "mock-model"},
            "prompt_template": {"system": "describe"},
            "gui_pipeline": {
                "source_sheet": "创建记录",
                "review_sheet": "审核结果",
            },
        },
    )

    calls = {}
    monkeypatch.setattr(gui_app, "ensure_pipeline_columns", lambda workbook, source_sheet: calls.update({"source_sheet": source_sheet}))
    monkeypatch.setattr(
        gui_app,
        "build_case_manifests",
        lambda workbook, source_sheet, allowed_statuses, local_root, server_root, mode: calls.update(
            {
                "allowed_statuses": allowed_statuses,
                "local_root": local_root,
                "server_root": server_root,
                "mode": mode,
            }
        )
        or [],
    )

    workbook = tmp_path / "records.xlsx"
    workbook.write_text("placeholder", encoding="utf-8")
    gui_app.launch_case_pipeline_gui(workbook_path=str(workbook))
    captured["scan_cases"]()

    assert calls["source_sheet"] == "创建记录"
    assert calls["allowed_statuses"] == gui_app.DEFAULT_ALLOWED_STATUSES
    assert calls["local_root"] == gui_app.DEFAULT_LOCAL_ROOT
    assert calls["server_root"] == gui_app.DEFAULT_SERVER_ROOT
    assert calls["mode"] == gui_app.DEFAULT_MODE


# ── 以下测试验证新 launch_case_pipeline_gui (Task 9) ─────────────────────────


def test_new_launch_loads_config_and_tag_options(monkeypatch, tmp_path: Path):
    """新 launch 函数加载 config.json 和 tag_options.json 并传给 MainWindow。"""
    import json
    from video_tagging_assistant.gui import app as gui_app

    config_data = {
        "workbook_path": str(tmp_path / "records.xlsx"),
        "mode": "OV50H40_Action5Pro_DCG HDR",
        "adb_exe": "adb.exe",
        "dut_root": "/mnt",
        "local_case_root": str(tmp_path),
        "server_upload_root": str(tmp_path / "server"),
        "intermediate_dir": str(tmp_path / "intermediate"),
        "provider": {"name": "mock", "model": "mock-model"},
        "prompt_template": {"system": "describe"},
    }
    tag_options_data = {
        "安装方式": ["手持"],
        "运动模式": ["行走"],
        "运镜方式": ["推U摇"],
        "光源": ["正常"],
        "画面特征": ["边缘"],
        "影像表达": ["风景录像"],
    }
    (tmp_path / "config.json").write_text(json.dumps(config_data), encoding="utf-8")
    (tmp_path / "tag_options.json").write_text(
        json.dumps(tag_options_data), encoding="utf-8"
    )

    captured = {}

    class FakeMainWindow:
        def __init__(self, config, tag_options):
            captured["config"] = config
            captured["tag_options"] = tag_options

        def show(self):
            captured["shown"] = True

    class FakeApp:
        @staticmethod
        def instance():
            return None

        def __init__(self, argv):
            pass

        def exec_(self):
            return 0

    monkeypatch.setattr(gui_app, "QApplication", FakeApp)
    monkeypatch.setattr(gui_app, "MainWindow", FakeMainWindow)
    monkeypatch.setattr(
        gui_app,
        "_CONFIG_PATH",
        tmp_path / "config.json",
    )
    monkeypatch.setattr(
        gui_app,
        "_TAG_OPTIONS_PATH",
        tmp_path / "tag_options.json",
    )

    result = gui_app.launch_case_pipeline_gui()

    assert captured["config"]["mode"] == "OV50H40_Action5Pro_DCG HDR"
    assert captured["tag_options"]["安装方式"] == ["手持"]
    assert captured.get("shown") is True
    assert result == 0
