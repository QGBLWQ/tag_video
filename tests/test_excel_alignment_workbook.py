from pathlib import Path

import openpyxl
import pytest

from video_tagging_assistant.excel_workbook import (
    clear_rk_raw_value,
    load_aligned_rk_raw_rows,
    load_rk_raw_values,
    write_rk_raw_value,
)


def _build_get_list_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "获取列表"
    ws.append(["日期", "20260508", "", ""])
    ws.append(["处理状态", "RK_raw", "Action5Pro_Nomal", "Action5Pro_Night"])
    ws.append(["R", "", "DJI_0001.MP4", "DJI_0101.MP4"])
    ws.append(["R", "32", "DJI_0002.MP4", "DJI_0102.MP4"])
    ws.append(["R", "32x", "DJI_0003.MP4", "DJI_0103.MP4"])
    wb.save(path)


def test_load_rk_raw_values_returns_rows_with_dji_values(tmp_path: Path):
    workbook_path = tmp_path / "alignment.xlsx"
    _build_get_list_workbook(workbook_path)

    assert load_rk_raw_values(workbook_path, source_sheet="获取列表") == {
        3: "",
        4: "32",
        5: "32x",
    }


def test_load_aligned_rk_raw_rows_returns_only_non_empty_rk_raw_rows(tmp_path: Path):
    workbook_path = tmp_path / "alignment.xlsx"
    _build_get_list_workbook(workbook_path)

    assert load_aligned_rk_raw_rows(workbook_path, source_sheet="获取列表") == {
        4: "32",
        5: "32x",
    }


def test_write_and_clear_rk_raw_values_update_target_rows_in_xlsx(tmp_path: Path):
    workbook_path = tmp_path / "alignment.xlsx"
    _build_get_list_workbook(workbook_path)

    write_rk_raw_value(workbook_path, source_sheet="获取列表", row_index=3, rk_raw_value="31")
    clear_rk_raw_value(workbook_path, source_sheet="获取列表", row_index=4)

    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["获取列表"]
    assert ws["B3"].value == "31"
    assert ws["B4"].value in ("", None)
    assert ws["B5"].value == "32x"


def test_write_rk_raw_value_rejects_xlsm(tmp_path: Path):
    workbook_path = tmp_path / "alignment.xlsm"
    workbook_path.write_bytes(b"fake xlsm")

    with pytest.raises(ValueError, match="xlsm"):
        write_rk_raw_value(workbook_path, source_sheet="获取列表", row_index=3, rk_raw_value="31")
