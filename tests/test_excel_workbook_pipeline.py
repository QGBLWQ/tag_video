from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from video_tagging_assistant.excel_workbook import (
    build_case_manifests,
    ensure_pipeline_columns,
    load_approved_review_rows,
    load_pipeline_cases,
    update_pipeline_status,
)


PIPELINE_HEADERS = [
    "序号",
    "文件夹名",
    "备注",
    "创建日期",
    "Raw存放路径",
    "VS_Nomal",
    "VS_Night",
    "安装方式",
    "运动模式",
]

REVIEW_HEADERS = [
    "文件夹名",
    "创建记录行号",
    "Raw存放路径",
    "视频路径",
    "自动简介",
    "自动标签",
    "自动画面描述",
    "审核结论",
    "人工修订简介",
    "人工修订标签",
    "审核备注",
    "审核人",
    "审核时间",
    "同步状态",
    "归档状态",
    "归档目标路径",
]

GET_LIST_HEADERS = ["处理状态", "RK_raw", "Action5Pro_Nomal", "Action5Pro_Night"]


def build_pipeline_workbook(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "创建记录"
    ws.append(PIPELINE_HEADERS)
    ws.append([
        1,
        "case_A_0105",
        "场景备注",
        "20260428",
        r"E:\DV\case_A_0105\case_A_0105_RK_raw_117",
        r"E:\DV\case_A_0105\case_A_0105_DJI_normal.MP4",
        r"E:\DV\case_A_0105\case_A_0105_night_DJI.MP4",
        "手持",
        "行走",
    ])
    review = wb.create_sheet("审核结果")
    review.append(REVIEW_HEADERS)
    wb.save(path)


def build_bridge_workbook(path: Path):
    wb = Workbook()
    create_record = wb.active
    create_record.title = "创建记录"
    create_record.append(PIPELINE_HEADERS)
    create_record.append([
        1,
        "case_A_0001",
        "场景备注",
        "20260422",
        r"E:\DV\case_A_0001\case_A_0001_RK_raw_117",
        r"E:\DV\case_A_0001\DJI_20260422151829_0001_D.MP4",
        r"E:\DV\case_A_0001\DJI_20260422151916_0021_D.MP4",
        "手持",
        "行走",
    ])
    get_list = wb.create_sheet("获取列表")
    get_list.append(["日期", "20260422", "", ""])
    get_list.append(GET_LIST_HEADERS)
    get_list.append(["R", "117", "DJI_20260422151829_0001_D.MP4", "DJI_20260422151916_0021_D.MP4"])
    review = wb.create_sheet("审核结果")
    review.append(REVIEW_HEADERS)
    wb.save(path)



def test_ensure_pipeline_columns_rejects_xlsm_workbook(tmp_path: Path):
    workbook_path = tmp_path / "records.xlsm"
    build_pipeline_workbook(workbook_path)

    with pytest.raises(ValueError) as exc:
        ensure_pipeline_columns(workbook_path, source_sheet="创建记录")

    assert ".xlsm" in str(exc.value)
    assert "只读" in str(exc.value)



def test_ensure_pipeline_columns_appends_runtime_headers(tmp_path: Path):
    workbook_path = tmp_path / "records.xlsx"
    build_pipeline_workbook(workbook_path)

    ensure_pipeline_columns(workbook_path, source_sheet="创建记录")

    wb = load_workbook(workbook_path)
    ws = wb["创建记录"]
    headers = [cell.value for cell in ws[1]]
    assert "pipeline_status" in headers
    assert "tag_status" in headers
    assert "updated_at" in headers



def test_build_case_manifests_from_get_list_bridges_back_to_create_record(tmp_path: Path):
    workbook_path = tmp_path / "records.xlsx"
    build_bridge_workbook(workbook_path)

    manifests = build_case_manifests(
        workbook_path,
        source_sheet="获取列表",
        allowed_statuses={"queued"},
        local_root=tmp_path / "local",
        server_root=tmp_path / "server",
        mode="OV50H40_Action5Pro_DCG HDR",
    )

    assert len(manifests) == 1
    assert manifests[0].case_id == "case_A_0001"
    assert manifests[0].created_date == "20260422"
    assert manifests[0].vs_normal_path.name == "DJI_20260422151829_0001_D.MP4"



def test_build_case_manifests_from_get_list_raises_when_create_record_match_missing(tmp_path: Path):
    workbook_path = tmp_path / "records.xlsx"
    build_bridge_workbook(workbook_path)

    wb = load_workbook(workbook_path)
    ws = wb["获取列表"]
    ws.cell(3, 2).value = "999"
    wb.save(workbook_path)

    with pytest.raises(ValueError) as exc:
        build_case_manifests(
            workbook_path,
            source_sheet="获取列表",
            allowed_statuses={"queued", ""},
            local_root=tmp_path / "local",
            server_root=tmp_path / "server",
            mode="OV50H40_Action5Pro_DCG HDR",
        )

    assert "RK_raw=999" in str(exc.value)
    assert "No matching create-record row found" in str(exc.value)



def test_build_case_manifests_from_get_list_raises_when_create_record_match_is_ambiguous(tmp_path: Path):
    workbook_path = tmp_path / "records.xlsx"
    build_bridge_workbook(workbook_path)

    wb = load_workbook(workbook_path)
    ws = wb["创建记录"]
    ws.append([
        2,
        "case_A_0002",
        "重复候选",
        "20260422",
        r"E:\DV\case_A_0002\case_A_0002_RK_raw_117",
        r"E:\DV\case_A_0002\DJI_20260422151829_0001_D.MP4",
        r"E:\DV\case_A_0002\DJI_20260422151916_0021_D.MP4",
        "手持",
        "行走",
    ])
    wb.save(workbook_path)

    with pytest.raises(ValueError) as exc:
        build_case_manifests(
            workbook_path,
            source_sheet="获取列表",
            allowed_statuses={"queued", ""},
            local_root=tmp_path / "local",
            server_root=tmp_path / "server",
            mode="OV50H40_Action5Pro_DCG HDR",
        )

    assert "Matched 2 create-record rows" in str(exc.value)
    assert "RK_raw=117" in str(exc.value)


def test_load_pipeline_cases_returns_only_pending_rows(tmp_path: Path):
    workbook_path = tmp_path / "records.xlsx"
    build_pipeline_workbook(workbook_path)
    ensure_pipeline_columns(workbook_path, source_sheet="创建记录")
    update_pipeline_status(
        workbook_path,
        source_sheet="创建记录",
        case_id="case_A_0105",
        status_updates={"pipeline_status": "queued"},
    )

    rows = load_pipeline_cases(workbook_path, source_sheet="创建记录", allowed_statuses={"queued"})
    assert [row.case_id for row in rows] == ["case_A_0105"]


def test_update_pipeline_status_writes_multiple_runtime_fields(tmp_path: Path):
    workbook_path = tmp_path / "records.xlsx"
    build_pipeline_workbook(workbook_path)
    ensure_pipeline_columns(workbook_path, source_sheet="创建记录")

    update_pipeline_status(
        workbook_path,
        source_sheet="创建记录",
        case_id="case_A_0105",
        status_updates={
            "pipeline_status": "completed",
            "pull_status": "done",
            "upload_status": "done",
            "last_error": "",
        },
    )

    wb = load_workbook(workbook_path)
    ws = wb["创建记录"]
    headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    assert ws.cell(2, headers["pipeline_status"]).value == "completed"
    assert ws.cell(2, headers["pull_status"]).value == "done"


def test_build_case_manifests_maps_excel_rows_to_runtime_paths(tmp_path: Path):
    workbook_path = tmp_path / "records.xlsx"
    build_pipeline_workbook(workbook_path)
    ensure_pipeline_columns(workbook_path, source_sheet="创建记录")
    update_pipeline_status(
        workbook_path,
        source_sheet="创建记录",
        case_id="case_A_0105",
        status_updates={"pipeline_status": "queued"},
    )

    manifests = build_case_manifests(
        workbook_path,
        source_sheet="创建记录",
        allowed_statuses={"queued"},
        local_root=tmp_path / "local",
        server_root=tmp_path / "server",
        mode="OV50H40_Action5Pro_DCG HDR",
    )

    assert len(manifests) == 1
    assert manifests[0].case_id == "case_A_0105"
    assert manifests[0].local_case_root == tmp_path / "local" / "OV50H40_Action5Pro_DCG HDR" / "20260428" / "case_A_0105"


def test_load_approved_review_rows_reads_excel_decisions(tmp_path: Path):
    workbook_path = tmp_path / "records.xlsx"
    build_pipeline_workbook(workbook_path)

    wb = load_workbook(workbook_path)
    ws = wb["审核结果"]
    ws.append([
        "case_A_0105",
        2,
        r"E:\DV\case_A_0105\case_A_0105_RK_raw_117",
        r"E:\DV\case_A_0105\case_A_0105_DJI_normal.MP4",
        "自动简介",
        "安装方式=手持",
        "自动画面描述",
        "修改后通过",
        "人工简介",
        "安装方式=肩扛",
        "补充备注",
        "tester",
        "2026-04-29 10:00:00",
        "",
        "",
        "",
    ])
    wb.save(workbook_path)

    rows = load_approved_review_rows(workbook_path, review_sheet="审核结果")

    assert len(rows) == 1
    assert rows[0]["case_id"] == "case_A_0105"
    assert rows[0]["review_decision"] == "修改后通过"
    assert rows[0]["manual_summary"] == "人工简介"
