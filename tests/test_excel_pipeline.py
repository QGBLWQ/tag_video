from pathlib import Path

from openpyxl import Workbook, load_workbook

from video_tagging_assistant.context_builder import build_prompt_context
from video_tagging_assistant.excel_models import ReviewSheetRow
from video_tagging_assistant.excel_workbook import load_confirmed_cases, sync_approved_rows, upsert_review_rows
from video_tagging_assistant.models import CompressedArtifact, GenerationResult, VideoTask


class StubProvider:
    provider_name = "stub"

    def generate(self, context):
        return GenerationResult(
            source_video_path=context.source_video_path,
            case_key=context.prompt_payload["workbook"]["文件夹名"],
            summary_text="自动简介",
            structured_tags={"安装方式": "手持", "运动模式": "行走"},
            scene_description="画面描述",
            provider="stub",
            model="stub-model",
        )


def build_source_workbook(path: Path, video_path: Path) -> None:
    wb = Workbook()
    source = wb.active
    source.title = "创建记录"
    source.append([
        "序号",
        "文件夹名",
        "备注",
        "Raw存放路径",
        "VS_Nomal",
        "VS_Night",
        "安装方式",
        "运动模式",
        "标签审核状态",
        "最终简介",
        "最终标签",
    ])
    source.append([
        1,
        "case_A_0001",
        "场景备注",
        "raw/path",
        str(video_path),
        "videos/case_A_0001/night.mp4",
        "手持",
        "行走",
        "待生成",
        "",
        "",
    ])
    wb.save(path)


def test_quick_excel_review_flow(tmp_path: Path):
    workbook_path = tmp_path / "records.xlsx"
    video_path = tmp_path / "videos" / "case_A_0001" / "clip01.mp4"
    video_path.parent.mkdir(parents=True)
    video_path.write_bytes(b"video")
    build_source_workbook(workbook_path, video_path)

    case_row = load_confirmed_cases(
        workbook_path,
        source_sheet="创建记录",
        case_key_column="文件夹名",
        status_column="标签审核状态",
    )[0]
    task = VideoTask(
        source_video_path=video_path,
        relative_path=Path("DCG_HDR") / case_row.case_key / video_path.name,
        file_name=video_path.name,
        case_id=case_row.case_key,
        mode="DCG_HDR",
    )
    artifact = CompressedArtifact(
        source_video_path=video_path,
        compressed_video_path=tmp_path / "output" / "compressed" / "clip01_proxy.mp4",
    )
    provider = StubProvider()
    result = provider.generate(build_prompt_context(task, artifact, {"system": "describe"}, case_row=case_row))

    upsert_review_rows(
        workbook_path,
        review_sheet="标签审核",
        rows=[
            ReviewSheetRow(
                case_key=case_row.case_key,
                workbook_row_index=case_row.workbook_row_index,
                raw_path=case_row.raw_path,
                video_path=case_row.vs_normal_path,
                auto_summary=result.summary_text,
                auto_tags=";".join(f"{k}={v}" for k, v in result.structured_tags.items()),
                auto_scene_description=result.scene_description,
                review_decision="审核通过",
            )
        ],
    )
    sync_approved_rows(workbook_path, source_sheet="创建记录", review_sheet="标签审核")

    wb = load_workbook(workbook_path)
    review_sheet = wb["标签审核"]
    source_sheet = wb["创建记录"]
    assert review_sheet["A2"].value == "case_A_0001"
    assert source_sheet["I2"].value == "审核通过"
    assert source_sheet["J2"].value == "自动简介"
    assert source_sheet["K2"].value == "安装方式=手持;运动模式=行走"
