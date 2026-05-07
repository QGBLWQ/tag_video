from pathlib import Path
from typing import List

import openpyxl
import pytest

from video_tagging_assistant.excel_workbook import TagResult, upsert_create_record_row
from video_tagging_assistant.pipeline_models import CaseManifest


CREATE_RECORD_SHEET = "\u521b\u5efa\u8bb0\u5f55"
CREATE_RECORD_HEADERS = [
    "\u5e8f\u53f7",
    "\u6587\u4ef6\u5939\u540d",
    "\u5907\u6ce8",
    "\u521b\u5efa\u65e5\u671f",
    "Null",
    "\u6570\u91cf",
    "\u5b89\u88c5\u65b9\u5f0f",
    "\u8fd0\u52a8\u6a21\u5f0f",
    "\u8fd0\u955c\u5143\u7d20",
    "\u5149\u6e90\u5212\u5206",
    "\u753b\u9762\u7279\u5f81",
    "\u5f71\u50cf\u8868\u8fbe",
    "Raw\u5b58\u653e\u8def\u5f84",
    "\u8bbe\u5907\u7f16\u53f7",
    "\u6a21\u7ec4\u578b\u53f7",
    "\u82af\u7247",
    "\u91c7\u96c6\u6a21\u5f0f",
    "bit\u4f4d",
    "\u5e27\u7387",
    "\u5176\u4ed6\u4fe1\u606f",
    "VS_Nomal",
    "VS_Night",
]


def _make_manifest(tmp_path: Path, case_id: str = "case_A_0001") -> CaseManifest:
    return CaseManifest(
        case_id=case_id,
        row_index=2,
        created_date="20260422",
        mode="OV50H40_Action5Pro_DCG HDR",
        raw_path=Path("/mnt/nvme/CapturedData/117"),
        vs_normal_path=Path("DJI_20260422151829_0001_D.MP4"),
        vs_night_path=Path("DJI_20260422151916_0021_D.MP4"),
        local_case_root=tmp_path / "cases" / "OV50H40_Action5Pro_DCG HDR" / "20260422" / case_id,
        server_case_dir=tmp_path / "server" / "OV50H40_Action5Pro_DCG HDR" / "20260422" / case_id,
        remark="",
    )


def _make_tag_result() -> TagResult:
    return TagResult(
        install_method="\u624b\u6301",
        motion_mode="\u884c\u8d70",
        camera_move="\u8def\u6d4b\u57fa\u51c6",
        light_source="\u6b63\u5e38",
        image_feature="\u8fb9\u7f18\u7279\u5f81 \u5f3a\u5f31",
        image_expression="\u98ce\u666f\u5f55\u50cf",
        review_status="\u5ba1\u6838\u901a\u8fc7",
        scene_description="\u65b0\u5907\u6ce8",
    )


def _build_workbook(path: Path, rows: List[List[object]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = CREATE_RECORD_SHEET
    ws.append(CREATE_RECORD_HEADERS)
    for row in rows:
        ws.append(row)
    wb.save(path)


def test_upsert_create_record_row_updates_existing_case_id_in_place(tmp_path: Path):
    wb_path = tmp_path / "create_record.xlsx"
    _build_workbook(
        wb_path,
        [
            [
                1,
                "case_A_0001",
                "old note",
                "20260421",
                "",
                "1",
                "",
                "",
                "",
                "",
                "",
                "",
                "old_raw",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        ],
    )

    manifest = _make_manifest(tmp_path)
    tag_result = _make_tag_result()

    upsert_create_record_row(wb_path, manifest, tag_result)

    wb = openpyxl.load_workbook(wb_path)
    ws = wb[CREATE_RECORD_SHEET]
    assert ws.max_row == 2
    headers = {cell.value: cell.column for cell in ws[1]}
    row = ws[2]
    assert row[headers["\u6587\u4ef6\u5939\u540d"] - 1].value == manifest.case_id
    assert row[headers["\u5907\u6ce8"] - 1].value == tag_result.scene_description
    assert row[headers["\u521b\u5efa\u65e5\u671f"] - 1].value == manifest.created_date
    assert row[headers["\u5b89\u88c5\u65b9\u5f0f"] - 1].value == tag_result.install_method
    assert row[headers["\u8fd0\u52a8\u6a21\u5f0f"] - 1].value == tag_result.motion_mode
    assert row[headers["\u8fd0\u955c\u5143\u7d20"] - 1].value == tag_result.camera_move
    assert row[headers["\u5149\u6e90\u5212\u5206"] - 1].value == tag_result.light_source
    assert row[headers["\u753b\u9762\u7279\u5f81"] - 1].value == tag_result.image_feature
    assert row[headers["\u5f71\u50cf\u8868\u8fbe"] - 1].value == tag_result.image_expression


def test_upsert_create_record_row_appends_new_row_for_current_xlsx_data(tmp_path: Path):
    wb_path = tmp_path / "create_record.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet"
    wb.save(wb_path)

    manifest = _make_manifest(tmp_path)
    tag_result = _make_tag_result()

    upsert_create_record_row(wb_path, manifest, tag_result)

    wb = openpyxl.load_workbook(wb_path)
    ws = wb[CREATE_RECORD_SHEET]
    assert ws.max_row == 2
    headers = {cell.value: cell.column for cell in ws[1]}
    assert ws.cell(2, headers["\u6587\u4ef6\u5939\u540d"]).value == manifest.case_id
    assert ws.cell(2, headers["Raw\u5b58\u653e\u8def\u5f84"]).value.endswith("case_A_0001_RK_raw_117")
    assert ws.cell(2, headers["VS_Nomal"]).value.endswith("case_A_0001_DJI_20260422151829_0001_D.MP4")
    assert ws.cell(2, headers["VS_Night"]).value.endswith("case_A_0001_night_DJI_20260422151916_0021_D.MP4")


def test_upsert_create_record_row_same_case_id_twice_does_not_append_duplicate_row(tmp_path: Path):
    wb_path = tmp_path / "create_record.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet"
    wb.save(wb_path)

    manifest = _make_manifest(tmp_path)
    first_result = _make_tag_result()
    second_result = TagResult(
        install_method="\u7a7f\u6234",
        motion_mode="\u8dd1\u52a8",
        camera_move="\u4f4e\u89d2\u5ea6\u8ddf\u62cd",
        light_source="\u9006\u5149",
        image_feature="\u5927\u9762\u79ef\u9ad8\u4eae",
        image_expression="\u57ce\u5e02\u8857\u666f",
        review_status="\u5ba1\u6838\u901a\u8fc7",
        scene_description="\u7b2c\u4e8c\u6b21\u5199\u56de",
    )

    upsert_create_record_row(wb_path, manifest, first_result)
    upsert_create_record_row(wb_path, manifest, second_result)

    wb = openpyxl.load_workbook(wb_path)
    ws = wb[CREATE_RECORD_SHEET]
    headers = {cell.value: cell.column for cell in ws[1]}

    assert ws.max_row == 2
    assert ws.cell(2, headers["\u6587\u4ef6\u5939\u540d"]).value == manifest.case_id
    assert ws.cell(2, headers["\u5907\u6ce8"]).value == second_result.scene_description
    assert ws.cell(2, headers["\u5b89\u88c5\u65b9\u5f0f"]).value == second_result.install_method
    assert ws.cell(2, headers["\u8fd0\u52a8\u6a21\u5f0f"]).value == second_result.motion_mode
    assert ws.cell(2, headers["\u8fd0\u955c\u5143\u7d20"]).value == second_result.camera_move
    assert ws.cell(2, headers["\u5149\u6e90\u5212\u5206"]).value == second_result.light_source


def test_upsert_create_record_row_sparse_existing_sheet_updates_available_columns(tmp_path: Path):
    wb_path = tmp_path / "sparse_create_record.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = CREATE_RECORD_SHEET
    ws.append(["\u5e8f\u53f7", "\u6587\u4ef6\u5939\u540d", "\u5907\u6ce8", "\u5b89\u88c5\u65b9\u5f0f"])
    ws.append([1, "case_A_0001", "old note", "old install"])
    wb.save(wb_path)

    manifest = _make_manifest(tmp_path)
    tag_result = _make_tag_result()

    upsert_create_record_row(wb_path, manifest, tag_result)

    wb = openpyxl.load_workbook(wb_path)
    ws = wb[CREATE_RECORD_SHEET]
    headers = {cell.value: cell.column for cell in ws[1]}

    assert ws.max_row == 2
    assert ws.cell(2, headers["\u6587\u4ef6\u5939\u540d"]).value == manifest.case_id
    assert ws.cell(2, headers["\u5907\u6ce8"]).value == tag_result.scene_description
    assert ws.cell(2, headers["\u5b89\u88c5\u65b9\u5f0f"]).value == tag_result.install_method


def test_upsert_create_record_row_rejects_xlsm(tmp_path: Path):
    xlsm_path = tmp_path / "create_record.xlsm"
    xlsm_path.write_bytes(b"fake xlsm")

    with pytest.raises(ValueError, match="xlsm"):
        upsert_create_record_row(xlsm_path, _make_manifest(tmp_path), _make_tag_result())
