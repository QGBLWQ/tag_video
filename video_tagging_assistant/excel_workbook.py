from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Set

from openpyxl import load_workbook

from video_tagging_assistant.excel_models import ConfirmedCaseRow, ReviewSheetRow
from video_tagging_assistant.pipeline_models import CaseManifest, ExcelCaseRecord

REVIEW_HEADERS = [
    "文件夹名",
    "创建记录行号",
    "Raw存放路径",
    "视频路径",
    "自动简介",
    "自动标签",
    "自动画面描述",
    "审核结论",
    "人工修订简介",
    "人工修订标签",
    "审核备注",
    "审核人",
    "审核时间",
    "同步状态",
    "归档状态",
    "归档目标路径",
]

PIPELINE_RUNTIME_HEADERS = [
    "pipeline_status",
    "tag_status",
    "review_status",
    "pull_status",
    "copy_status",
    "upload_status",
    "last_error",
    "run_id",
    "updated_at",
]

GET_LIST_REQUIRED_HEADERS = {"处理状态", "RK_raw", "Action5Pro_Nomal", "Action5Pro_Night"}


@dataclass
class GetListRow:
    created_date: str
    status: str
    rk_raw: str
    vs_normal_name: str
    vs_night_name: str


def _reject_xlsm_write(workbook_path: Path) -> None:
    if Path(workbook_path).suffix.lower() == ".xlsm":
        raise ValueError(f"{workbook_path} 是 .xlsm 工作簿，当前仅支持只读访问，禁止原地写回以避免损坏台账")


def _header_map(sheet) -> Dict[str, int]:
    return {str(cell.value).strip(): idx + 1 for idx, cell in enumerate(sheet[1]) if cell.value is not None}


def _header_map_for_row(sheet, row_index: int) -> Dict[str, int]:
    return {
        str(cell.value).strip(): idx + 1
        for idx, cell in enumerate(sheet[row_index])
        if cell.value is not None
    }


def _extract_raw_suffix(raw_path: str) -> str:
    return Path(raw_path).name.split("_")[-1]


def _load_get_list_rows(workbook_path: Path, source_sheet: str) -> List[GetListRow]:
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook[source_sheet]
    created_date = str(sheet.cell(1, 2).value or "").strip()
    headers = _header_map_for_row(sheet, 2)
    missing = GET_LIST_REQUIRED_HEADERS - set(headers)
    if missing:
        raise ValueError(f"获取列表 缺少必要表头: {sorted(missing)}")

    rows: List[GetListRow] = []
    for row_index in range(3, sheet.max_row + 1):
        rk_raw = str(sheet.cell(row_index, headers["RK_raw"]).value or "").strip()
        normal = str(sheet.cell(row_index, headers["Action5Pro_Nomal"]).value or "").strip()
        night = str(sheet.cell(row_index, headers["Action5Pro_Night"]).value or "").strip()
        if not rk_raw and not normal and not night:
            continue
        rows.append(
            GetListRow(
                created_date=created_date,
                status=str(sheet.cell(row_index, headers["处理状态"]).value or "").strip(),
                rk_raw=rk_raw,
                vs_normal_name=normal,
                vs_night_name=night,
            )
        )
    return rows


def _match_create_record_rows(create_record_rows: List[ExcelCaseRecord], get_list_row: GetListRow) -> ExcelCaseRecord:
    matches = [
        row
        for row in create_record_rows
        if _extract_raw_suffix(row.raw_path) == get_list_row.rk_raw
        and Path(row.vs_normal_path).name == get_list_row.vs_normal_name
        and Path(row.vs_night_path).name == get_list_row.vs_night_name
    ]
    if not matches:
        raise ValueError(
            "No matching create-record row found for "
            f"RK_raw={get_list_row.rk_raw}, normal={get_list_row.vs_normal_name}, night={get_list_row.vs_night_name}"
        )
    if len(matches) > 1:
        raise ValueError(
            "Matched "
            f"{len(matches)} create-record rows for RK_raw={get_list_row.rk_raw}, "
            f"normal={get_list_row.vs_normal_name}, night={get_list_row.vs_night_name}"
        )
    return matches[0]


def ensure_pipeline_columns(workbook_path: Path, source_sheet: str) -> None:
    _reject_xlsm_write(workbook_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook[source_sheet]
    headers = _header_map(sheet)
    next_column = sheet.max_column + 1
    for header in PIPELINE_RUNTIME_HEADERS:
        if header not in headers:
            sheet.cell(1, next_column).value = header
            next_column += 1
    workbook.save(workbook_path)


def _load_create_record_rows(workbook_path: Path, source_sheet: str) -> List[ExcelCaseRecord]:
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook[source_sheet]
    headers = _header_map(sheet)
    rows: List[ExcelCaseRecord] = []
    for row_index in range(2, sheet.max_row + 1):
        case_id = str(sheet.cell(row_index, headers["文件夹名"]).value or "").strip()
        if not case_id:
            continue
        rows.append(
            ExcelCaseRecord(
                row_index=row_index,
                case_id=case_id,
                created_date=str(sheet.cell(row_index, headers["创建日期"]).value or "").strip(),
                remark=str(sheet.cell(row_index, headers["备注"]).value or "").strip(),
                raw_path=str(sheet.cell(row_index, headers["Raw存放路径"]).value or "").strip(),
                vs_normal_path=str(sheet.cell(row_index, headers["VS_Nomal"]).value or "").strip(),
                vs_night_path=str(sheet.cell(row_index, headers["VS_Night"]).value or "").strip(),
                labels={
                    "安装方式": str(sheet.cell(row_index, headers["安装方式"]).value or "").strip(),
                    "运动模式": str(sheet.cell(row_index, headers["运动模式"]).value or "").strip(),
                },
                pipeline_status="",
            )
        )
    return rows


def load_pipeline_cases(workbook_path: Path, source_sheet: str, allowed_statuses: Set[str]) -> List[ExcelCaseRecord]:
    workbook = load_workbook(workbook_path)
    sheet = workbook[source_sheet]
    headers = _header_map(sheet)
    rows: List[ExcelCaseRecord] = []
    for row_index in range(2, sheet.max_row + 1):
        case_id = str(sheet.cell(row_index, headers["文件夹名"]).value or "").strip()
        if not case_id:
            continue
        status = str(sheet.cell(row_index, headers["pipeline_status"]).value or "").strip()
        if status not in allowed_statuses:
            continue
        rows.append(
            ExcelCaseRecord(
                row_index=row_index,
                case_id=case_id,
                created_date=str(sheet.cell(row_index, headers["创建日期"]).value or "").strip(),
                remark=str(sheet.cell(row_index, headers["备注"]).value or "").strip(),
                raw_path=str(sheet.cell(row_index, headers["Raw存放路径"]).value or "").strip(),
                vs_normal_path=str(sheet.cell(row_index, headers["VS_Nomal"]).value or "").strip(),
                vs_night_path=str(sheet.cell(row_index, headers["VS_Night"]).value or "").strip(),
                labels={
                    "安装方式": str(sheet.cell(row_index, headers["安装方式"]).value or "").strip(),
                    "运动模式": str(sheet.cell(row_index, headers["运动模式"]).value or "").strip(),
                },
                pipeline_status=status,
            )
        )
    return rows


def update_pipeline_status(workbook_path: Path, source_sheet: str, case_id: str, status_updates: Dict[str, str]) -> None:
    _reject_xlsm_write(workbook_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook[source_sheet]
    headers = _header_map(sheet)
    for row_index in range(2, sheet.max_row + 1):
        current_case_id = str(sheet.cell(row_index, headers["文件夹名"]).value or "").strip()
        if current_case_id != case_id:
            continue
        for key, value in status_updates.items():
            sheet.cell(row_index, headers[key]).value = value
        break
    workbook.save(workbook_path)


def build_case_manifests(
    workbook_path: Path,
    source_sheet: str,
    allowed_statuses: Set[str],
    local_root: Path,
    server_root: Path,
    mode: str,
) -> List[CaseManifest]:
    if source_sheet == "获取列表":
        create_record_rows = _load_create_record_rows(
            workbook_path,
            source_sheet="创建记录",
        )
        rows = [
            _match_create_record_rows(create_record_rows, row)
            for row in _load_get_list_rows(workbook_path, source_sheet)
        ]
    else:
        rows = load_pipeline_cases(workbook_path, source_sheet=source_sheet, allowed_statuses=allowed_statuses)

    manifests: List[CaseManifest] = []
    for row in rows:
        manifests.append(
            CaseManifest(
                case_id=row.case_id,
                row_index=row.row_index,
                created_date=row.created_date,
                mode=mode,
                raw_path=Path(row.raw_path),
                vs_normal_path=Path(row.vs_normal_path),
                vs_night_path=Path(row.vs_night_path),
                local_case_root=Path(local_root) / mode / row.created_date / row.case_id,
                server_case_dir=Path(server_root) / mode / row.created_date / row.case_id,
                remark=row.remark,
                labels=row.labels,
            )
        )
    return manifests


def get_next_case_sequence(workbook_path: Path, pc_id: str) -> int:
    """Read 「创建记录」 and return the next available sequence number for pc_id.

    Returns 1 if sheet doesn't exist or has no matching rows.
    """
    if not workbook_path.exists():
        return 1
    workbook = load_workbook(workbook_path, data_only=True)
    if "创建记录" not in workbook.sheetnames:
        return 1
    sheet = workbook["创建记录"]
    headers = _header_map(sheet)
    if "文件夹名" not in headers:
        return 1
    prefix = f"case_{pc_id}_"
    max_seq = 0
    for row_index in range(2, sheet.max_row + 1):
        case_id = str(sheet.cell(row_index, headers["文件夹名"]).value or "").strip()
        if case_id.startswith(prefix):
            try:
                seq = int(case_id[len(prefix):])
                if seq > max_seq:
                    max_seq = seq
            except ValueError:
                pass
    return max_seq + 1


def load_get_list_manifests(
    workbook_path: Path,
    source_sheet: str,
    pc_id: str,
    dji_nomal_dir: Path,
    dji_night_dir: Path,
    local_root: Path,
    server_root: Path,
    mode: str,
    starting_sequence: int = 1,
) -> List[CaseManifest]:
    """Read 「获取列表」 only and build CaseManifest objects.

    Does NOT require 「创建记录」 to exist. case_id is derived as case_{pc_id}_{seq:04d}.
    Designed for the tagging tab where 「创建记录」 is populated only after review approval.
    """
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook[source_sheet]
    created_date = str(sheet.cell(1, 2).value or "").strip()
    headers = _header_map_for_row(sheet, 2)
    missing = GET_LIST_REQUIRED_HEADERS - set(headers)
    if missing:
        raise ValueError(f"获取列表 缺少必要表头: {sorted(missing)}")

    manifests: List[CaseManifest] = []
    sequence = starting_sequence
    for row_index in range(3, sheet.max_row + 1):
        rk_raw = str(sheet.cell(row_index, headers["RK_raw"]).value or "").strip()
        normal = str(sheet.cell(row_index, headers["Action5Pro_Nomal"]).value or "").strip()
        night = str(sheet.cell(row_index, headers["Action5Pro_Night"]).value or "").strip()
        if not rk_raw and not normal and not night:
            continue
        case_id = f"case_{pc_id}_{sequence:04d}"
        manifests.append(
            CaseManifest(
                case_id=case_id,
                row_index=row_index,
                created_date=created_date,
                mode=mode,
                raw_path=Path(rk_raw),
                vs_normal_path=Path(dji_nomal_dir) / normal if normal else Path(normal),
                vs_night_path=Path(dji_night_dir) / night if night else Path(night),
                local_case_root=Path(local_root) / mode / created_date / case_id,
                server_case_dir=Path(server_root) / mode / created_date / case_id,
                remark="",
                labels={},
            )
        )
        sequence += 1
    return manifests


def load_confirmed_cases(
    workbook_path: Path,
    source_sheet: str,
    case_key_column: str,
    status_column: str,
) -> List[ConfirmedCaseRow]:
    workbook = load_workbook(workbook_path)
    sheet = workbook[source_sheet]
    headers = _header_map(sheet)

    rows: List[ConfirmedCaseRow] = []
    for row_index in range(2, sheet.max_row + 1):
        case_value = sheet.cell(row_index, headers[case_key_column]).value
        if not case_value:
            continue
        rows.append(
            ConfirmedCaseRow(
                case_key=str(case_value).strip(),
                workbook_row_index=row_index,
                raw_path=str(sheet.cell(row_index, headers["Raw存放路径"]).value or "").strip(),
                vs_normal_path=str(sheet.cell(row_index, headers["VS_Nomal"]).value or "").strip(),
                vs_night_path=str(sheet.cell(row_index, headers["VS_Night"]).value or "").strip(),
                note=str(sheet.cell(row_index, headers["备注"]).value or "").strip(),
                attributes={
                    "安装方式": str(sheet.cell(row_index, headers["安装方式"]).value or "").strip(),
                    "运动模式": str(sheet.cell(row_index, headers["运动模式"]).value or "").strip(),
                },
            )
        )
    return rows


def upsert_review_rows(workbook_path: Path, review_sheet: str, rows: List[ReviewSheetRow]) -> None:
    _reject_xlsm_write(workbook_path)
    workbook = load_workbook(workbook_path)
    if review_sheet in workbook.sheetnames:
        sheet = workbook[review_sheet]
    else:
        sheet = workbook.create_sheet(review_sheet)
        sheet.append(REVIEW_HEADERS)

    headers = _header_map(sheet)
    existing = {}
    for row_index in range(2, sheet.max_row + 1):
        case_key = sheet.cell(row_index, headers["文件夹名"]).value
        if case_key:
            existing[str(case_key).strip()] = row_index

    for row in rows:
        row_index = existing.get(row.case_key, sheet.max_row + 1)
        values = {
            "文件夹名": row.case_key,
            "创建记录行号": row.workbook_row_index,
            "Raw存放路径": row.raw_path,
            "视频路径": row.video_path,
            "自动简介": row.auto_summary,
            "自动标签": row.auto_tags,
            "自动画面描述": row.auto_scene_description,
            "审核结论": row.review_decision,
            "人工修订简介": row.manual_summary,
            "人工修订标签": row.manual_tags,
            "审核备注": row.review_note,
            "审核人": row.reviewer,
            "审核时间": row.reviewed_at,
            "同步状态": row.sync_status,
            "归档状态": row.archive_status,
            "归档目标路径": row.archive_target_path,
        }
        for header, value in values.items():
            sheet.cell(row_index, headers[header]).value = value

    workbook.save(workbook_path)


def load_approved_review_rows(workbook_path: Path, review_sheet: str) -> List[Dict[str, str]]:
    workbook = load_workbook(workbook_path)
    sheet = workbook[review_sheet]
    headers = _header_map(sheet)
    rows: List[Dict[str, str]] = []
    for row_index in range(2, sheet.max_row + 1):
        decision = str(sheet.cell(row_index, headers["审核结论"]).value or "").strip()
        if decision not in {"审核通过", "修改后通过"}:
            continue
        rows.append(
            {
                "case_id": str(sheet.cell(row_index, headers["文件夹名"]).value or "").strip(),
                "review_decision": decision,
                "manual_summary": str(sheet.cell(row_index, headers["人工修订简介"]).value or "").strip(),
                "manual_tags": str(sheet.cell(row_index, headers["人工修订标签"]).value or "").strip(),
                "review_note": str(sheet.cell(row_index, headers["审核备注"]).value or "").strip(),
            }
        )
    return rows


def sync_approved_rows(workbook_path: Path, source_sheet: str, review_sheet: str) -> None:
    _reject_xlsm_write(workbook_path)
    workbook = load_workbook(workbook_path)
    source = workbook[source_sheet]
    review = workbook[review_sheet]
    source_headers = _header_map(source)
    review_headers = _header_map(review)

    for row_index in range(2, review.max_row + 1):
        decision = str(review.cell(row_index, review_headers["审核结论"]).value or "").strip()
        if decision not in {"审核通过", "修改后通过"}:
            continue
        source_row = int(review.cell(row_index, review_headers["创建记录行号"]).value)
        auto_summary = str(review.cell(row_index, review_headers["自动简介"]).value or "").strip()
        auto_tags = str(review.cell(row_index, review_headers["自动标签"]).value or "").strip()
        manual_summary = str(review.cell(row_index, review_headers["人工修订简介"]).value or "").strip()
        manual_tags = str(review.cell(row_index, review_headers["人工修订标签"]).value or "").strip()

        source.cell(source_row, source_headers["标签审核状态"]).value = decision
        source.cell(source_row, source_headers["最终简介"]).value = manual_summary or auto_summary
        source.cell(source_row, source_headers["最终标签"]).value = manual_tags or auto_tags

    workbook.save(workbook_path)


@dataclass
class TagResult:
    """审核通过后，人工确认的完整标签结果。"""
    install_method: str    # 安装方式（单选）
    motion_mode: str       # 运动模式（单选）
    camera_move: str       # 运镜元素（单选）
    light_source: str      # 光源划分（单选）
    image_feature: str     # 画面特征（从 AI 多选中人工选一）
    image_expression: str  # 影像表达（从 AI 多选中人工选一）
    review_status: str     # 固定值 "审核通过"
    scene_description: str = ""                          # 画面描述（写入备注列）
    device_info: Dict[str, str] = field(default_factory=dict)  # 设备编号/芯片等


_CREATE_RECORD_HEADERS = [
    "序号", "文件夹名", "备注", "创建日期", "Null", "数量",
    "安装方式", "运动模式", "运镜元素", "光源划分", "画面特征", "影像表达",
    "Raw存放路径", "设备编号", "模组型号", "芯片", "采集模式", "bit位", "帧率", "其他信息",
    "VS_Nomal", "VS_Night",
]


def load_dut_info(workbook_path: Path) -> List[Dict[str, str]]:
    """Read Dut_info sheet. Returns list of device dicts; default device (默认选项=是) first."""
    workbook = load_workbook(workbook_path, data_only=True)
    if "Dut_info" not in workbook.sheetnames:
        return []
    sheet = workbook["Dut_info"]
    headers = _header_map(sheet)
    dut_fields = ["设备编号", "模组型号", "芯片", "采集模式", "bit位", "帧率", "其他信息"]
    devices: List[Dict[str, str]] = []
    default_device = None
    for row_index in range(2, sheet.max_row + 1):
        device_id_col = headers.get("设备编号", 2)
        device_id = str(sheet.cell(row_index, device_id_col).value or "").strip()
        if not device_id:
            continue
        device: Dict[str, str] = {}
        for col_name in dut_fields:
            if col_name in headers:
                device[col_name] = str(sheet.cell(row_index, headers[col_name]).value or "").strip()
        is_default = str(sheet.cell(row_index, headers.get("默认选项", 1)).value or "").strip() == "是"
        if is_default:
            default_device = device
        else:
            devices.append(device)
    if default_device is not None:
        devices.insert(0, default_device)
    return devices


def upsert_create_record_row(
    workbook_path: Path,
    manifest: "CaseManifest",
    tag_result: TagResult,
) -> None:
    """审核通过后，在「创建记录」sheet 末尾追加一行。sheet 不存在时自动创建并写入表头。

    workbook_path 必须是 .xlsx。
    """
    _reject_xlsm_write(workbook_path)
    workbook = load_workbook(workbook_path)

    if "创建记录" not in workbook.sheetnames:
        sheet = workbook.create_sheet("创建记录")
        sheet.append(_CREATE_RECORD_HEADERS)
    else:
        sheet = workbook["创建记录"]

    headers = _header_map(sheet)
    target_row = sheet.max_row + 1

    # 构建服务器路径（Windows UNC 格式）
    server_case = str(manifest.server_case_dir).replace("/", "\\")
    case_id = manifest.case_id
    vs_nomal = f"{server_case}\\{case_id}_{manifest.vs_normal_path.name}"
    vs_night = f"{server_case}\\{case_id}_night_{manifest.vs_night_path.name}"
    raw_path = f"{server_case}\\{case_id}_RK_raw_{manifest.raw_path.name}"

    data: Dict[str, str] = {
        "序号": str(target_row - 1),
        "文件夹名": case_id,
        "备注": tag_result.scene_description,
        "创建日期": manifest.created_date,
        "数量": "1",
        "安装方式": tag_result.install_method,
        "运动模式": tag_result.motion_mode,
        "运镜元素": tag_result.camera_move,
        "光源划分": tag_result.light_source,
        "画面特征": tag_result.image_feature,
        "影像表达": tag_result.image_expression,
        "Raw存放路径": raw_path,
        "VS_Nomal": vs_nomal,
        "VS_Night": vs_night,
        **tag_result.device_info,
    }
    for col_name, value in data.items():
        if col_name in headers:
            sheet.cell(row=target_row, column=headers[col_name]).value = value

    workbook.save(workbook_path)
